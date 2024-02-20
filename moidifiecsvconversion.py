import pandas as pd
import os
from openpyxl import load_workbook

def convert_xlsx_to_csv(input_folder, output_folder):
    if not os.path.isdir(input_folder):
        raise ValueError(f"Input folder '{input_folder}' does not exist.")

    if not os.path.isdir(output_folder):
        raise ValueError(f"Output folder '{output_folder}' does not exist.")

    for filename in os.listdir(input_folder):
        if filename.endswith(".xlsx"):
            try:
                # Load the Excel file using openpyxl
                wb = load_workbook(os.path.join(input_folder, filename))

                # Disable filters in all sheets
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    ws.auto_filter.ref = None

                # Save the modified Excel file
                wb.save(os.path.join(input_folder, filename))

                # Read the modified Excel file using pandas
                df = pd.read_excel(os.path.join(input_folder, filename))

                # Generate a descriptive output filename (optional)
                output_filename = os.path.splitext(filename)[0] + ".csv"

                # Save the DataFrame to CSV, addressing potential issues
                df.to_csv(os.path.join(output_folder, output_filename), index=False, header=True)

                print(f"Converted '{filename}' to '{output_filename}'")
            except Exception as e:
                print(f"Error converting '{filename}': {e}")

if __name__ == "__main__":
    input_folder = r"D:\Try"  # Use raw string to handle backslashes
    output_folder = r"D:\Try"

    convert_xlsx_to_csv(input_folder, output_folder)
